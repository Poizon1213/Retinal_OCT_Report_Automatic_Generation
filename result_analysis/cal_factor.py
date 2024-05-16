# -*- coding:utf-8 -*-
from save import *
import re
from string import punctuation
import openpyxl
import pandas as pd
import os
import numpy as np
from sklearn.metrics import roc_auc_score, accuracy_score, precision_score, recall_score, f1_score


'''处理实验结果到excel'''
data_folder = r'E:\学习\毕业论文\data\20220110\1124'
# with open(os.path.join(data_folder, 'pred.txt'), 'r', encoding='utf-8') as f:
#     preddata = f.readlines()
#     f.close()
#
# with open(os.path.join(data_folder, 'target.txt'), 'r', encoding='utf-8') as f:
#     targetdata = f.readlines()
#     f.close()
#
# patho_classify_file = openpyxl.load_workbook(os.path.join(data_folder, 'result_181920.xlsx'))
# sheet = patho_classify_file['Pathological_Info']
#
# small_categories = {}
# classcol = 2
#
# for i in range(len(preddata)):
#     checknum = preddata[i].strip().split()[0]
#     if len(preddata[i].strip().split()) == 1:
#         pred_diagnosis = ''
#     else:
#         pred_diagnosis = preddata[i].strip().split()[1].split('。')[:-1]
#     target_diagnosis = targetdata[i].strip().split()[1].split('。')[:-1]
#
#     sheet.cell(i*2+2, 1, checknum)
#     sheet.cell(i*2+3, 1, checknum)
#
#     for seq in pred_diagnosis:
#         structure_disease = ''
#         if seq in disease_to_struc:
#             structure_disease = disease_to_struc[seq]
#         else:
#             if '全层裂孔' in seq:
#                 structure_disease += '组织缺失'
#             if 'RPE层脱离灶下' in seq:
#                 structure_disease += 'RPE层脱离灶下信号'
#             if '增殖膜' in seq:
#                 structure_disease += '视网膜前增殖膜'
#             if 'RPE层高反射小团点' in seq:
#                 structure_disease += 'RPE层反射强弱'
#             if '神经上皮层内层组织反射' in seq:
#                 structure_disease += '分层组织反射异常'
#             if '黄斑<unk>结构未见明显异常' in seq or '未见明显正常中心凹结构' in seq:
#                 structure_disease += '中心凹结构'
#         if structure_disease != '':
#             if structure_disease not in small_categories:
#                 small_categories[structure_disease] = classcol
#                 sheet.cell(1, small_categories[structure_disease], structure_disease)
#                 classcol += 1
#             pre_seq = sheet.cell(i*2 + 2, small_categories[structure_disease]).value
#             if type(pre_seq) == str:
#                 sheet.cell(i*2 + 2, small_categories[structure_disease], pre_seq + seq + '。')
#             else:
#                 sheet.cell(i*2 + 2, small_categories[structure_disease], seq + '。')
#         else:
#             print('pred_diagnosis', checknum, seq)
#
#     for seq in target_diagnosis:
#         structure_disease = ''
#         if seq in disease_to_struc:
#             structure_disease = disease_to_struc[seq]
#         else:
#             if '全层裂孔' in seq:
#                 structure_disease += '组织缺失'
#             if 'RPE层脱离灶下' in seq:
#                 structure_disease += 'RPE层脱离灶下信号'
#             if '增殖膜' in seq:
#                 structure_disease += '视网膜前增殖膜'
#             if 'RPE层高反射小团点' in seq:
#                 structure_disease += 'RPE层反射强弱'
#             if '神经上皮层内层组织反射' in seq:
#                 structure_disease += '分层组织反射异常'
#             if '黄斑<unk>结构未见明显异常' in seq or '未见明显正常中心凹结构' in seq:
#                 structure_disease += '中心凹结构'
#         if structure_disease != '':
#             if structure_disease not in small_categories:
#                 small_categories[structure_disease] = classcol
#                 sheet.cell(1, small_categories[structure_disease], structure_disease)
#                 classcol += 1
#             pre_seq = sheet.cell(i * 2 + 3, small_categories[structure_disease]).value
#             if type(pre_seq) == str:
#                 sheet.cell(i * 2 + 3, small_categories[structure_disease], pre_seq + seq + '。')
#             else:
#                 sheet.cell(i * 2 + 3, small_categories[structure_disease], seq + '。')
#         else:
#             print('target_diagnosis', checknum, seq)
#
# patho_classify_file.save(os.path.join(data_folder, 'result_181920.xlsx'))
#
# print(small_categories)

'''计算指标'''
# df_struct_rpe_ref = pd.read_excel(os.path.join(data_folder, 'result_181920.xlsx'), sheet_name='Pathological_Info', header=None).values.tolist()
#
# print(df_struct_rpe_ref[0])
# factors = {}
#
# for col in range(1, 38):
#     factors[df_struct_rpe_ref[0][col]] = {'TP': 0, 'FP': 0, 'FN': 0, 'TN': 0, 'amount': 0, 'truth_label': [], 'pred_label': []}
# print(factors)
# print(len(factors))
# y_truth = []
# y_pred = []
#
# y_seq_truth = []
# y_seq_pred = []
#
# seq_amount = dict()
# for i in range(1, len(df_struct_rpe_ref)-2, 2):
#     y_single_truth = [0 for id in range(38)]
#     y_single_pred = [0 for id in range(38)]
#     for col in range(1, 38):
#         # TP
#         if type(df_struct_rpe_ref[i][col]) == str and type(df_struct_rpe_ref[i+1][col]) == str:
#             factors[df_struct_rpe_ref[0][col]]['TP'] += 1
#         # FP
#         elif type(df_struct_rpe_ref[i][col]) == str and type(df_struct_rpe_ref[i+1][col]) != str:
#             factors[df_struct_rpe_ref[0][col]]['FP'] += 1
#         # FN
#         elif type(df_struct_rpe_ref[i][col]) != str and type(df_struct_rpe_ref[i+1][col]) == str:
#             factors[df_struct_rpe_ref[0][col]]['FN'] += 1
#         # TN
#         elif type(df_struct_rpe_ref[i][col]) != str and type(df_struct_rpe_ref[i+1][col]) != str:
#             factors[df_struct_rpe_ref[0][col]]['TN'] += 1
#
#         if type(df_struct_rpe_ref[i+1][col]) == str:
#             factors[df_struct_rpe_ref[0][col]]['amount'] += 1
#             y_single_truth[col-1] = 1
#             factors[df_struct_rpe_ref[0][col]]['truth_label'].append(1)
#         else:
#             factors[df_struct_rpe_ref[0][col]]['truth_label'].append(0)
#
#         if type(df_struct_rpe_ref[i][col]) == str:
#             y_single_pred[col-1] = 1
#             factors[df_struct_rpe_ref[0][col]]['pred_label'].append(1)
#         else:
#             factors[df_struct_rpe_ref[0][col]]['pred_label'].append(0)
#
#     y_single_seq_truth = [0 for id in range(131)]
#     y_single_seq_pred = [0 for id in range(131)]
#
#     for col in range(1, 38):
#         if type(df_struct_rpe_ref[i + 1][col]) == str:
#             for s in df_struct_rpe_ref[i + 1][col].split('。')[:-1]:
#                 if s not in seq_num:
#                     if '衰减' in s:
#                         y_single_seq_truth[seq_num['RPE层脱离灶下信号递减']] = 1
#                         if 'RPE层脱离灶下信号递减' not in seq_amount:
#                             seq_amount['RPE层脱离灶下信号递减'] = 0
#                         seq_amount['RPE层脱离灶下信号递减'] += 1
#                     elif '呈' in s:
#                         y_single_seq_truth[seq_num[s[:8] + s[9:]]] = 1
#                         if s[:8] + s[9:] not in seq_amount:
#                             seq_amount[s[:8] + s[9:]] = 0
#                         seq_amount[s[:8] + s[9:]] += 1
#                     else:
#                         print(s)
#                 else:
#                     y_single_seq_truth[seq_num[s]] = 1
#                     if s not in seq_amount:
#                         seq_amount[s] = 0
#                     seq_amount[s] += 1
#
#         if type(df_struct_rpe_ref[i][col]) == str:
#             for s in df_struct_rpe_ref[i][col].split('。')[:-1]:
#                 if s not in seq_num:
#                     if '衰减' in s:
#                         y_single_seq_pred[seq_num['RPE层脱离灶下信号递减']] = 1
#                     elif '呈' in s:
#                         y_single_seq_pred[seq_num[s[:8] + s[9:]]] = 1
#                     else:
#                         print(s)
#                 else:
#                     y_single_seq_pred[seq_num[s]] = 1
#
#     y_truth.append(y_single_truth)
#     y_pred.append(y_single_pred)
#
#     y_seq_pred.append(y_single_seq_pred)
#     y_seq_truth.append(y_single_seq_truth)
#
# for i in range(38):
#     if 1 not in y_truth[:][i]:
#         print("find only one class")
#
#
# # y_truth = np.array(y_truth)
# # y_pred = np.array(y_pred)
#
#
# # print(factors)
# accuracy = {}
# precision = {}
# recall = {}
#
# accuracylib = {}
# precisionlib = {}
# recalllib = {}
# print(factors)
# for disease in factors:
#     accuracy[disease] = (factors[disease]['TP'] + factors[disease]['TN']) / (factors[disease]['TP'] + factors[disease]['TN'] + factors[disease]['FP'] + factors[disease]['FN'])
#     if (factors[disease]['TP'] + factors[disease]['FP']) == 0:
#         precision[disease] = 0
#     else:
#         precision[disease] = (factors[disease]['TP']) / (factors[disease]['TP'] + factors[disease]['FP'])
#     if (factors[disease]['TP'] + factors[disease]['FN']) == 0:
#         recall[disease] = 0
#     else:
#         recall[disease] = (factors[disease]['TP']) / (factors[disease]['TP'] + factors[disease]['FN'])
#     accuracylib[disease] = accuracy_score(factors[disease]['truth_label'], factors[disease]['pred_label'])
#     precisionlib[disease] = precision_score(factors[disease]['truth_label'], factors[disease]['pred_label'])
#     recalllib[disease] = recall_score(factors[disease]['truth_label'], factors[disease]['pred_label'])
#
# accuracylib_seq = dict()
# precisionlib_seq = dict()
# recalllib_seq = dict()
#
# for i, disease in enumerate(seq_num):
#     if disease not in seq_amount:
#         seq_amount[disease] = 0
#     # print([y_truth[j][i] for j in range(len(y_truth))])
#     # print([y_pred[j][i] for j in range(len(y_pred))])
#     accuracylib_seq[disease] = accuracy_score([y_seq_truth[j][i] for j in range(len(y_seq_truth))], [y_seq_pred[j][i] for j in range(len(y_seq_pred))])
#     precisionlib_seq[disease] = precision_score([y_seq_truth[j][i] for j in range(len(y_seq_truth))], [y_seq_pred[j][i] for j in range(len(y_seq_pred))])
#     recalllib_seq[disease] = recall_score([y_seq_truth[j][i] for j in range(len(y_seq_truth))], [y_seq_pred[j][i] for j in range(len(y_seq_pred))])
#
# print(accuracy)
# print(accuracylib)
# print(precision)
# print(precisionlib)
# print(recall)
# print(recalllib)
#
# print('accuracylib_seq', accuracylib_seq)
# print('precisionlib_seq', precisionlib_seq)
# print('recalllib_seq', recalllib_seq)
#
# # print('auc: ', roc_auc_score(y_truth, y_pred, average='micro'))
#
# patho_classify_file = openpyxl.load_workbook(os.path.join(data_folder, 'indicator.xlsx'))
# sheet = patho_classify_file['Sheet1']
#
# sheet.cell(2, 1, 'accuracy')
# sheet.cell(3, 1, 'precision')
# sheet.cell(4, 1, 'recall')
# sheet.cell(5, 1, 'amount')
#
# for col, disease in enumerate(factors):
#     sheet.cell(1, col+2, disease)
#     sheet.cell(2, col+2, accuracy[disease])
#     sheet.cell(3, col+2, precision[disease])
#     sheet.cell(4, col+2, recall[disease])
#     sheet.cell(5, col+2, factors[disease]['amount'])
#
# patho_classify_file.save(os.path.join(data_folder, 'indicator.xlsx'))
#
# patho_classify_file = openpyxl.load_workbook(os.path.join(data_folder, 'indicator_seq.xlsx'))
# sheet = patho_classify_file['Sheet1']
#
# sheet.cell(2, 1, 'accuracy')
# sheet.cell(3, 1, 'precision')
# sheet.cell(4, 1, 'recall')
# sheet.cell(5, 1, 'amount')
#
# for col, disease in enumerate(seq_num):
#     sheet.cell(1, col+2, disease)
#     sheet.cell(2, col+2, accuracylib_seq[disease])
#     sheet.cell(3, col+2, precisionlib_seq[disease])
#     sheet.cell(4, col+2, recalllib_seq[disease])
#     sheet.cell(5, col+2, seq_amount[disease])
#
# patho_classify_file.save(os.path.join(data_folder, 'indicator_seq.xlsx'))

'''计算指标（分小句）'''
df_struct_rpe_ref = pd.read_excel(os.path.join(data_folder, 'result_181920.xlsx'), sheet_name='Pathological_Info', header=None).values.tolist()
factor = dict()
for seq in seq_num:
    factor[seq] = {'TP': 0, 'TN': 0, 'FP': 0, 'FN': 0, 'amount': 0}
for i in range(1, len(df_struct_rpe_ref)-2, 2):
    appreance = []
    for col in range(1, 38):
        seq_truth = []
        seq_pred = []
        if type(df_struct_rpe_ref[i + 1][col]) == str:
            seq_truth = df_struct_rpe_ref[i + 1][col].split('。')[:-1]
        if type(df_struct_rpe_ref[i][col]) == str:
            seq_pred = df_struct_rpe_ref[i][col].split('。')[:-1]
        for seq in seq_truth:
            if seq in seq_pred:
                factor[seq]['TP'] += 1
            else:
                factor[seq]['FN'] += 1
            appreance.append(seq)
            factor[seq]['amount'] += 1
        for seq in seq_pred:
            if seq not in seq_truth:
                factor[seq]['FP'] += 1
                appreance.append(seq)
    for seq in factor:
        if seq not in appreance:
            factor[seq]['TN'] += 1


patho_classify_file = openpyxl.load_workbook(os.path.join(data_folder, 'indicator_single_FNTP.xlsx'))
sheet = patho_classify_file['Sheet1']

sheet.cell(2, 1, 'TP')
sheet.cell(3, 1, 'TN')
sheet.cell(4, 1, 'FN')
sheet.cell(5, 1, 'FP')
sheet.cell(6, 1, 'amount')

for col, disease in enumerate(factor):
    sheet.cell(1, col+2, disease)
    sheet.cell(2, col+2, factor[disease]['TP'])
    sheet.cell(3, col+2, factor[disease]['TN'])
    sheet.cell(4, col+2, factor[disease]['FN'])
    sheet.cell(5, col+2, factor[disease]['FP'])
    sheet.cell(6, col+2, factor[disease]['amount'])

patho_classify_file.save(os.path.join(data_folder, 'indicator_single_FNTP.xlsx'))
